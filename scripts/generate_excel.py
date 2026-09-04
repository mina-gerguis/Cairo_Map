import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel():
    with open("scripts/processed_places.json", "r", encoding="utf-8") as f:
        places = json.load(f)

    print(f"Total places loaded: {len(places)}")

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active

    # Define Styles
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    sub_header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    sub_header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

    card_title_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    card_title_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    card_value_font = Font(name="Segoe UI", size=18, bold=True, color="1E3A8A")

    regular_font = Font(name="Segoe UI", size=10, color="0F172A")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    link_font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
    
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_fill_alt = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    gov_cairo_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    gov_giza_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="E2E8F0")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    card_border_side = Side(border_style="medium", color="CBD5E1")
    card_border = Border(left=card_border_side, right=card_border_side, top=card_border_side, bottom=card_border_side)

    headers = [
        ("م", 6),
        ("اسم المكان (بالعربية)", 34),
        ("اسم المكان (بالإنجليزية)", 30),
        ("المحافظة", 14),
        ("المنطقة / الحي", 24),
        ("التصنيف الرئيسي", 25),
        ("النوع والتفصيل", 24),
        ("العنوان بالتفصيل", 38),
        ("خط العرض (Lat)", 15),
        ("خط الطول (Lon)", 15),
        ("رقم الهاتف / التواصل", 20),
        ("مواعيد العمل", 22),
        ("الوصف والمعلومات", 35),
        ("رابط خرائط جوجل (Google Maps)", 26),
        ("الموقع الإلكتروني", 26)
    ]

    def format_sheet_headers(ws, title_headers):
        ws.views.sheetView[0].rightToLeft = True
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28
        
        for col_idx, (header_text, width) in enumerate(title_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = cell_border
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

    def populate_places(ws, place_list):
        format_sheet_headers(ws, headers)
        for row_idx, p in enumerate(place_list, start=2):
            ws.row_dimensions[row_idx].height = 22
            is_alt = (row_idx % 2 == 0)
            base_fill = row_fill_alt if is_alt else row_fill_white
            
            # Row cells
            # 1: ID
            c1 = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.font = bold_font
            c1.fill = base_fill
            c1.border = cell_border
            
            # 2: Arabic Name
            c2 = ws.cell(row=row_idx, column=2, value=p["name_ar"])
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c2.font = bold_font
            c2.fill = base_fill
            c2.border = cell_border
            
            # 3: English Name
            c3 = ws.cell(row=row_idx, column=3, value=p["name_en"])
            c3.alignment = Alignment(horizontal="left", vertical="center")
            c3.font = regular_font
            c3.fill = base_fill
            c3.border = cell_border
            
            # 4: Governorate
            c4 = ws.cell(row=row_idx, column=4, value=p["governorate"])
            c4.alignment = Alignment(horizontal="center", vertical="center")
            c4.font = bold_font
            c4.fill = gov_cairo_fill if p["governorate"] == "القاهرة" else gov_giza_fill
            c4.border = cell_border
            
            # 5: District
            c5 = ws.cell(row=row_idx, column=5, value=p["district"])
            c5.alignment = Alignment(horizontal="center", vertical="center")
            c5.font = regular_font
            c5.fill = base_fill
            c5.border = cell_border
            
            # 6: Main Category
            c6 = ws.cell(row=row_idx, column=6, value=p["main_category"])
            c6.alignment = Alignment(horizontal="center", vertical="center")
            c6.font = bold_font
            c6.fill = base_fill
            c6.border = cell_border
            
            # 7: Sub Category
            c7 = ws.cell(row=row_idx, column=7, value=p["sub_category"])
            c7.alignment = Alignment(horizontal="center", vertical="center")
            c7.font = regular_font
            c7.fill = base_fill
            c7.border = cell_border
            
            # 8: Address
            c8 = ws.cell(row=row_idx, column=8, value=p["address"])
            c8.alignment = Alignment(horizontal="right", vertical="center")
            c8.font = regular_font
            c8.fill = base_fill
            c8.border = cell_border
            
            # 9: Lat
            c9 = ws.cell(row=row_idx, column=9, value=p["latitude"])
            c9.alignment = Alignment(horizontal="center", vertical="center")
            c9.font = regular_font
            c9.fill = base_fill
            c9.border = cell_border
            
            # 10: Lon
            c10 = ws.cell(row=row_idx, column=10, value=p["longitude"])
            c10.alignment = Alignment(horizontal="center", vertical="center")
            c10.font = regular_font
            c10.fill = base_fill
            c10.border = cell_border
            
            # 11: Phone
            c11 = ws.cell(row=row_idx, column=11, value=p["phone"] if p["phone"] else "-")
            c11.alignment = Alignment(horizontal="center", vertical="center")
            c11.font = regular_font
            c11.fill = base_fill
            c11.border = cell_border
            
            # 12: Opening Hours
            c12 = ws.cell(row=row_idx, column=12, value=p["opening_hours"] if p["opening_hours"] else "-")
            c12.alignment = Alignment(horizontal="center", vertical="center")
            c12.font = regular_font
            c12.fill = base_fill
            c12.border = cell_border
            
            # 13: Description
            c13 = ws.cell(row=row_idx, column=13, value=p["description"])
            c13.alignment = Alignment(horizontal="right", vertical="center")
            c13.font = regular_font
            c13.fill = base_fill
            c13.border = cell_border
            
            # 14: Google Maps Link
            c14 = ws.cell(row=row_idx, column=14, value="فتح في Google Maps")
            c14.hyperlink = p["gmaps_url"]
            c14.alignment = Alignment(horizontal="center", vertical="center")
            c14.font = link_font
            c14.fill = base_fill
            c14.border = cell_border
            
            # 15: Website
            if p["website"]:
                c15 = ws.cell(row=row_idx, column=15, value="زيارة الموقع")
                web_url = p["website"] if p["website"].startswith("http") else f"https://{p['website']}"
                c15.hyperlink = web_url
                c15.alignment = Alignment(horizontal="center", vertical="center")
                c15.font = link_font
            else:
                c15 = ws.cell(row=row_idx, column=15, value="-")
                c15.alignment = Alignment(horizontal="center", vertical="center")
                c15.font = regular_font
            c15.fill = base_fill
            c15.border = cell_border

    # 1. Master Sheet: All Places
    ws_all = wb.create_sheet(title="دليل كافة الأماكن (الشامل)")
    populate_places(ws_all, places)

    # 2. Category sheets
    category_map = {
        "السياحة والآثار والمتاحف": "السياحة والآثار والمتاحف",
        "الفنادق والضيافة": "الفنادق والمنتجعات",
        "المولات ومراكز التسوق": "المولات ومراكز التسوق",
        "المستشفيات والمراكز الطبية": "المستشفيات والطبية",
        "الجامعات والتعليم": "الجامعات والتعليم",
        "المطاعم والكافيهات": "المطاعم والكافيهات",
        "الحدائق والنوادي والترفيه": "الحدائق والنوادي والترفيه",
        "النقل والمواصلات والمطارات": "النقل والمواصلات",
        "الأسواق والهايبر ماركت": "الأسواق والهايبرماركت",
        "السفارات والمباني الحكومية": "السفارات والمباني الحكومية"
    }

    category_counts = {}
    for p in places:
        c = p["main_category"]
        category_counts[c] = category_counts.get(c, 0) + 1

    for cat_key, sheet_title in category_map.items():
        cat_places = [p for p in places if p["main_category"] == cat_key]
        if cat_places:
            ws_cat = wb.create_sheet(title=sheet_title[:31])
            populate_places(ws_cat, cat_places)

    # 3. Top 100 Places Sheet (أهم 100 مكان مميز ومختار في القاهرة والجيزة)
    # Filter high-interest places across diverse categories
    top_100_places = []
    seen_top = set()
    
    # Priority order for top 100
    prio_cats = [
        "السياحة والآثار والمتاحف",
        "المولات ومراكز التسوق",
        "الفنادق والضيافة",
        "المستشفيات والمراكز الطبية",
        "الجامعات والتعليم",
        "الحدائق والنوادي والترفيه",
        "النقل والمواصلات والمطارات",
        "المطاعم والكافيهات"
    ]
    
    for cat in prio_cats:
        cat_items = [p for p in places if p["main_category"] == cat]
        # pick diverse items from both Cairo & Giza
        cairo_items = [p for p in cat_items if p["governorate"] == "القاهرة"]
        giza_items = [p for p in cat_items if p["governorate"] == "الجيزة"]
        
        # take top items
        for p in (cairo_items[:10] + giza_items[:6]):
            if p["name_ar"] not in seen_top and len(top_100_places) < 100:
                seen_top.add(p["name_ar"])
                top_100_places.append(p)
                
    # fill remaining up to 100 if any
    for p in places:
        if len(top_100_places) >= 100:
            break
        if p["name_ar"] not in seen_top:
            seen_top.add(p["name_ar"])
            top_100_places.append(p)

    ws_top100 = wb.create_sheet(title="أفضل 100 مكان مميز", index=0)
    populate_places(ws_top100, top_100_places)

    # 4. Summary & Dashboard Sheet
    ws_summary = wb.create_sheet(title="لوحة الإحصائيات والملخص", index=0)
    ws_summary.views.sheetView[0].rightToLeft = True
    ws_summary.showGridLines = True

    # Title Banner
    ws_summary.merge_cells("A1:G1")
    title_cell = ws_summary.cell(row=1, column=1, value="📊 الدليل الشامل لأماكن ومعالم القاهرة والجيزة - ملخص البيانات")
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    title_cell.fill = header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[1].height = 40

    # Summary KPI Cards
    total_count = len(places)
    cairo_count = sum(1 for p in places if p["governorate"] == "القاهرة")
    giza_count = sum(1 for p in places if p["governorate"] == "الجيزة")
    tourism_count = sum(1 for p in places if "السياحة" in p["main_category"])
    malls_count = sum(1 for p in places if "المولات" in p["main_category"])
    hosp_count = sum(1 for p in places if "المستشفيات" in p["main_category"])
    edu_count = sum(1 for p in places if "الجامعات" in p["main_category"])
    hotel_count = sum(1 for p in places if "الفنادق" in p["main_category"])

    cards = [
        ("إجمالي الأماكن الموثقة", f"{total_count:,}", "B3:C3", "B4:C4", "1E3A8A"),
        ("أماكن محافظة القاهرة", f"{cairo_count:,}", "D3:E3", "D4:E4", "0D9488"),
        ("أماكن محافظة الجيزة", f"{giza_count:,}", "F3:G3", "F4:G4", "D97706"),
    ]

    for title, val, r1, r2, color_hex in cards:
        ws_summary.merge_cells(r1)
        ws_summary.merge_cells(r2)
        top_left_1 = ws_summary[r1.split(":")[0]]
        top_left_2 = ws_summary[r2.split(":")[0]]
        
        top_left_1.value = title
        top_left_1.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        top_left_1.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        top_left_1.alignment = Alignment(horizontal="center", vertical="center")
        
        top_left_2.value = val
        top_left_2.font = Font(name="Segoe UI", size=20, bold=True, color=color_hex)
        top_left_2.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        top_left_2.alignment = Alignment(horizontal="center", vertical="center")
        
    ws_summary.row_dimensions[3].height = 24
    ws_summary.row_dimensions[4].height = 36

    # Category Breakdown Table
    ws_summary.cell(row=6, column=2, value="📁 توزيع الأماكن حسب التصنيف الرئيسي").font = Font(name="Segoe UI", size=12, bold=True, color="1E3A8A")
    
    ws_summary.cell(row=7, column=2, value="التصنيف").fill = sub_header_fill
    ws_summary.cell(row=7, column=2).font = sub_header_font
    ws_summary.cell(row=7, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    ws_summary.cell(row=7, column=3, value="عدد الأماكن").fill = sub_header_fill
    ws_summary.cell(row=7, column=3).font = sub_header_font
    ws_summary.cell(row=7, column=3).alignment = Alignment(horizontal="center", vertical="center")
    
    ws_summary.cell(row=7, column=4, value="النسبة").fill = sub_header_fill
    ws_summary.cell(row=7, column=4).font = sub_header_font
    ws_summary.cell(row=7, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws_summary.row_dimensions[7].height = 24

    sorted_cats = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
    cur_row = 8
    for cat_name, cnt in sorted_cats:
        ws_summary.row_dimensions[cur_row].height = 20
        fill_cur = row_fill_alt if cur_row % 2 == 0 else row_fill_white
        
        c_name = ws_summary.cell(row=cur_row, column=2, value=cat_name)
        c_name.font = bold_font
        c_name.alignment = Alignment(horizontal="right", vertical="center")
        c_name.fill = fill_cur
        c_name.border = cell_border
        
        c_cnt = ws_summary.cell(row=cur_row, column=3, value=cnt)
        c_cnt.font = regular_font
        c_cnt.alignment = Alignment(horizontal="center", vertical="center")
        c_cnt.fill = fill_cur
        c_cnt.border = cell_border
        
        pct = f"{(cnt / total_count) * 100:.1f}%"
        c_pct = ws_summary.cell(row=cur_row, column=4, value=pct)
        c_pct.font = regular_font
        c_pct.alignment = Alignment(horizontal="center", vertical="center")
        c_pct.fill = fill_cur
        c_pct.border = cell_border
        
        cur_row += 1

    # District Breakdown Table
    district_counts = {}
    for p in places:
        d = p["district"]
        district_counts[d] = district_counts.get(d, 0) + 1

    ws_summary.cell(row=6, column=6, value="📍 توزيع الأماكن حسب أهم المناطق والأحياء").font = Font(name="Segoe UI", size=12, bold=True, color="1E3A8A")
    
    ws_summary.cell(row=7, column=6, value="المنطقة / الحي").fill = sub_header_fill
    ws_summary.cell(row=7, column=6).font = sub_header_font
    ws_summary.cell(row=7, column=6).alignment = Alignment(horizontal="center", vertical="center")
    
    ws_summary.cell(row=7, column=7, value="عدد الأماكن").fill = sub_header_fill
    ws_summary.cell(row=7, column=7).font = sub_header_font
    ws_summary.cell(row=7, column=7).alignment = Alignment(horizontal="center", vertical="center")

    sorted_districts = sorted(district_counts.items(), key=lambda x: x[1], reverse=True)[:15]
    cur_d_row = 8
    for dist_name, cnt in sorted_districts:
        ws_summary.row_dimensions[cur_d_row].height = 20
        fill_cur = row_fill_alt if cur_d_row % 2 == 0 else row_fill_white
        
        d_name = ws_summary.cell(row=cur_d_row, column=6, value=dist_name)
        d_name.font = bold_font
        d_name.alignment = Alignment(horizontal="right", vertical="center")
        d_name.fill = fill_cur
        d_name.border = cell_border
        
        d_cnt = ws_summary.cell(row=cur_d_row, column=7, value=cnt)
        d_cnt.font = regular_font
        d_cnt.alignment = Alignment(horizontal="center", vertical="center")
        d_cnt.fill = fill_cur
        d_cnt.border = cell_border
        
        cur_d_row += 1

    # Adjust widths for summary sheet
    ws_summary.column_dimensions["A"].width = 5
    ws_summary.column_dimensions["B"].width = 30
    ws_summary.column_dimensions["C"].width = 16
    ws_summary.column_dimensions["D"].width = 14
    ws_summary.column_dimensions["E"].width = 6
    ws_summary.column_dimensions["F"].width = 32
    ws_summary.column_dimensions["G"].width = 16

    # Remove the initial default sheet
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    # Save outputs in both Arabic and English filenames
    output_ar = "دليل_أماكن_القاهرة_والجيزة_3500_مكان.xlsx"
    output_en = "Cairo_and_Giza_Places_Directory_3500.xlsx"
    
    wb.save(output_ar)
    wb.save(output_en)
    print("Files saved successfully!")

if __name__ == "__main__":
    build_excel()
