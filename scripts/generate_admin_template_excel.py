import json
import os
import shutil
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("scripts/raw_places.json", "r", encoding="utf-8") as f:
    raw_data = json.load(f)

print(f"Loaded raw items: {len(raw_data)}")

def format_names(tags):
    name_ar = tags.get("name:ar", "").strip()
    name_en = tags.get("name:en", "").strip()
    name_default = tags.get("name", "").strip()
    
    has_arabic = bool(re.search(r'[\u0600-\u06FF]', name_default))
    
    if not name_ar:
        if has_arabic:
            name_ar = name_default
        else:
            name_ar = name_default
            
    if not name_en:
        if not has_arabic and name_default:
            name_en = name_default
        else:
            name_en = tags.get("int_name", "") or tags.get("name:fr", "")
            
    if not name_ar and name_en:
        name_ar = name_en
        
    return name_ar.strip(), name_en.strip()

def map_to_admin_category(tags):
    tourism = tags.get("tourism")
    historic = tags.get("historic")
    amenity = tags.get("amenity")
    shop = tags.get("shop")
    leisure = tags.get("leisure")
    office = tags.get("office")

    # 1. Food & Drinks
    if amenity in ["restaurant"]:
        cuisine = tags.get("cuisine", "")
        if any(x in cuisine.lower() for x in ["burger", "pizza", "sandwich", "fast_food", "kfc", "mcdonalds"]):
            return "أكل ومشروبات", "فاست فود", "مطعم فاست فود", "bx-cheese"
        return "أكل ومشروبات", "مطاعم", "مطعم", "bx-restaurant"
    if amenity in ["fast_food"]:
        return "أكل ومشروبات", "فاست فود", "مطعم وجبات سريعة", "bx-cheese"
    if amenity in ["cafe"]:
        return "أكل ومشروبات", "كافيهات", "كافيه ومقهى", "bx-coffee"
    if shop in ["bakery", "pastry", "confectionery"]:
        return "أكل ومشروبات", "مخابز وحلويات", "حلواني ومخبز", "bx-cookie"
    if amenity in ["juice_bar", "ice_cream"]:
        return "أكل ومشروبات", "عصائر", "محل عصائر ومثلجات", "bx-drink"

    # 2. Health
    if amenity in ["hospital"]:
        return "صحة", "مستشفيات", "مستشفى", "bx-plus-medical"
    if amenity in ["clinic", "doctors"]:
        return "صحة", "عيادات", "عيادة ومركز طبي", "bx-first-aid"
    if amenity in ["pharmacy"]:
        return "صحة", "صيدليات", "صيدلية", "bx-capsule"
    if amenity in ["dentist"]:
        return "صحة", "عيادات أسنان", "عيادة أسنان", "bx-smile"
    if amenity in ["laboratory"]:
        return "صحة", "معامل تحاليل", "معمل تحاليل", "bx-test-tube"

    # 3. Shopping
    if shop in ["mall"]:
        return "تسوق", "مولات", "مول تجاري", "bx-store-alt"
    if shop in ["supermarket", "convenience", "hypermarket", "grocery"]:
        return "تسوق", "سوبر ماركت", "سوبرماركت", "bx-cart"
    if shop in ["mobile_phone", "electronics"]:
        return "تسوق", "محلات موبايلات", "متجر إلكترونيات وموبايلات", "bx-mobile"
    if shop in ["computer"]:
        return "تسوق", "محلات كمبيوتر", "متجر كمبيوتر وتقنية", "bx-laptop"
    if shop in ["clothes", "fashion", "boutique"]:
        return "تسوق", "ملابس", "محل ملابس", "bx-closet"
    if shop in ["shoes"]:
        return "تسوق", "أحذية", "محل أحذية", "bx-football"
    if shop in ["jewelry"]:
        return "تسوق", "مجوهرات", "محل مجوهرات ومصوغات", "bx-diamond"
    if shop in ["books", "stationery"]:
        return "تسوق", "مكتبات", "مكتبة وأدوات مدرسية", "bx-book"
    if shop in ["furniture"]:
        return "تسوق", "أثاث", "معرض أثاث وديكور", "bx-home"
    if shop in ["beauty", "cosmetics", "perfumery"]:
        return "تسوق", "مستحضرات تجميل", "محل تجميل وعطور", "bx-heart"
    if shop in ["toys"]:
        return "تسوق", "ألعاب", "محل ألعاب أطفال", "bx-joystick"
    if shop:
        return "تسوق", "مولات", "متجر تجاري", "bx-store"

    # 4. Tourism & Hospitality
    if tourism in ["hotel", "motel", "resort"]:
        return "إقامة وسياحة", "فنادق", "فندق", "bx-hotel"
    if tourism in ["apartment", "guest_house"]:
        return "إقامة وسياحة", "شقق فندقية", "شقق فندقية وإقامة", "bx-building"
    if tourism in ["travel_agency"] or office in ["travel_agent"]:
        return "إقامة وسياحة", "شركات سياحة", "شركة سياحة وسفر", "bx-paper-plane"

    # 5. Entertainment & Culture
    if tourism in ["museum", "gallery"] or historic in ["museum"]:
        return "ترفيه", "متاحف", "متحف ومعرض أثري", "bx-mask"
    if amenity in ["cinema"]:
        return "ترفيه", "سينما", "دار سينما", "bx-film"
    if amenity in ["theatre"]:
        return "ترفيه", "مسارح", "مسرح وقاعة عروض", "bx-mask"
    if tourism in ["theme_park", "attraction"] or leisure in ["water_park", "amusement_arcade"]:
        if leisure == "water_park":
            return "ترفيه", "أكوا بارك", "مدينة ألعاب مائية", "bx-swim"
        return "ترفيه", "ملاهي", "مدينة ملاهي وألعاب", "bx-laugh"
    if historic:
        return "ترفيه", "متاحف", "معلم تاريخي وأثري", "bx-buildings"

    # 6. Sports & Fitness
    if leisure in ["fitness_centre", "sports_centre"]:
        return "رياضة", "جيم", "نادي صحي وجيم", "bx-dumbbell"
    if leisure in ["pitch", "stadium", "track"]:
        return "رياضة", "ملاعب", "ملعب واستاد رياضي", "bx-football"
    if leisure in ["swimming_pool"]:
        return "ترفيه", "حمامات سباحة", "حمام سباحة", "bx-swim"

    # 7. Public Places & Parks
    if leisure in ["park", "garden"]:
        return "أماكن عامة", "حدائق", "حديقة عامة ومتنزه", "bx-tree"
    if tourism in ["viewpoint"]:
        return "أماكن عامة", "ميادين", "مطل ومعلم عام", "bx-map"

    # 8. Education
    if amenity in ["university"]:
        return "تعليم", "جامعات", "جامعة وصرح أكاديمي", "bx-book-reader"
    if amenity in ["college", "school"]:
        return "تعليم", "مدارس", "مدرسة ومعهد تعليمي", "bx-book"
    if amenity in ["training_centre", "language_school"]:
        return "تعليم", "مراكز تعليم", "مركز تدريب ولغات", "bx-chalkboard"

    # 9. Finance
    if amenity in ["bank"]:
        return "خدمات مالية", "بنوك", "فرع بنك ومصرف", "bx-credit-card-front"
    if amenity in ["atm"]:
        return "خدمات مالية", "ماكينات ATM", "ماكينة صراف آلي ATM", "bx-credit-card-front"
    if amenity in ["bureau_de_change"]:
        return "خدمات مالية", "صرافة", "شركة صرافة", "bx-transfer"

    # 10. Government & Diplomatic
    if amenity in ["embassy"]:
        return "خدمات حكومية", "مصالح حكومية", "سفارة وقنصلية", "bx-buildings"
    if amenity in ["police"]:
        return "خدمات حكومية", "قسم شرطة", "قسم شرطة ونقطة أمنية", "bx-shield"
    if amenity in ["courthouse"]:
        return "خدمات حكومية", "محاكم", "مجمع محاكم", "bx-briefcase"
    if amenity in ["post_office"]:
        return "خدمات حكومية", "بريد", "مكتب بريد مصري", "bx-envelope"
    if amenity in ["townhall"]:
        return "خدمات حكومية", "مصالح حكومية", "هيئة ومصلحة حكومية", "bx-buildings"

    # 11. Automotive
    if amenity in ["fuel"]:
        return "سيارات", "محطات بنزين", "محطة وقود وبنزينة", "bx-gas-pump"
    if shop in ["car_repair"] or amenity in ["car_repair"]:
        return "سيارات", "مراكز صيانة", "مركز صيانة وخدمة سيارات", "bx-cog"
    if amenity in ["car_wash"]:
        return "سيارات", "مغاسل سيارات", "مغسلة سيارات", "bx-water"
    if amenity in ["parking"]:
        return "سيارات", "مواقف سيارات", "موقف وساحة انتظار", "bx-car"

    # 12. Religion
    if amenity in ["place_of_worship"]:
        religion = tags.get("religion", "")
        if religion == "christian":
            return "أماكن دينية", "كنائس", "كنيسة وقبطية", "bx-bookmark-heart"
        return "أماكن دينية", "مساجد", "مسجد وجامع", "bx-bookmark-heart"

    # 13. Services & Business
    if office:
        return "أعمال", "شركات", "شركة ومقر أعمال", "bx-briefcase"
    if shop in ["laundry", "dry_cleaning"]:
        return "خدمات", "مغاسل", "مغسلة ودراي كلين", "bx-sun"
    if shop in ["hairdresser", "barber"]:
        return "خدمات", "صالونات حلاقة", "صالون حلاقة ورجالي", "bx-cut"

    return "خدمات", "شركات شحن", "مرفق وخدمات عامة", "bx-store"

def determine_coords(item):
    if "lat" in item and "lon" in item:
        return round(float(item["lat"]), 6), round(float(item["lon"]), 6)
    if "center" in item:
        return round(float(item["center"]["lat"]), 6), round(float(item["center"]["lon"]), 6)
    return None, None

def determine_district_and_gov(lat, lon, tags):
    district_tag = tags.get("addr:district", "") or tags.get("addr:suburb", "") or tags.get("addr:neighbourhood", "") or tags.get("addr:quarter", "")
    
    gov = "القاهرة"
    dist = "وسط البلد"

    if 30.80 <= lon <= 31.05:
        gov = "الجيزة"
        dist = "الشيخ زايد" if lat >= 30.00 else "مدينة 6 أكتوبر"
    elif 31.05 < lon <= 31.16:
        gov = "الجيزة"
        dist = "الهرم وفيصل" if lat <= 30.01 else "أكتوبر / كرداسة"
    elif 31.16 < lon <= 31.225:
        if 30.050 <= lat <= 30.075 and 31.214 <= lon <= 31.228:
            gov = "القاهرة"
            dist = "الزمالك"
        elif 30.035 <= lat <= 30.050 and 31.218 <= lon <= 31.229:
            gov = "القاهرة"
            dist = "جاردن سيتي"
        elif lat >= 30.07:
            gov = "الجيزة"
            dist = "إمبابة والوراق"
        elif 30.045 <= lat < 30.07:
            gov = "الجيزة"
            dist = "المهندسين والعجوزة"
        elif 30.02 <= lat < 30.045:
            gov = "الجيزة"
            dist = "الدقي"
        else:
            gov = "الجيزة"
            dist = "الجيزة والعمرانية"
    elif 31.225 < lon <= 31.27:
        gov = "القاهرة"
        if 30.035 <= lat <= 30.065:
            dist = "وسط البلد"
        elif 30.065 < lat <= 30.12:
            dist = "شبرا"
        elif 29.93 <= lat < 30.00:
            dist = "المعادي"
        elif 30.00 <= lat < 30.035:
            dist = "مصر القديمة والمنيل"
        elif lat < 29.93:
            dist = "حلوان"
        else:
            dist = "العباسية والوايلي"
    elif 31.27 < lon <= 31.36:
        gov = "القاهرة"
        if 30.07 <= lat <= 30.15:
            dist = "مصر الجديدة"
        elif 30.02 <= lat < 30.07:
            dist = "مدينة نصر"
        elif 29.98 <= lat < 30.02:
            dist = "المقطم"
        elif 29.93 <= lat < 29.98:
            dist = "المعادي الجديدة"
        elif lat < 29.93:
            dist = "حلوان"
        else:
            dist = "عين شمس والمطرية"
    elif 31.36 < lon <= 31.48:
        gov = "القاهرة"
        if 30.08 <= lat <= 30.18:
            dist = "النزهة ومطار القاهرة"
        elif 29.96 <= lat < 30.08:
            dist = "التجمع الخامس"
        else:
            dist = "القاهرة الجديدة"
    elif lon > 31.48:
        gov = "القاهرة"
        if lat >= 30.10:
            dist = "مدينة الشروق"
        elif lat >= 30.03:
            dist = "مدينتي والرحاب"
        else:
            dist = "العاصمة الإدارية والتجمع الأول"

    if district_tag:
        dt = district_tag.strip()
        if any(x in dt for x in ["Maadi", "معادي"]):
            dist = "المعادي"
            gov = "القاهرة"
        elif any(x in dt for x in ["Nasr", "نصر"]):
            dist = "مدينة نصر"
            gov = "القاهرة"
        elif any(x in dt for x in ["Heliopolis", "مصر الجديدة"]):
            dist = "مصر الجديدة"
            gov = "القاهرة"
        elif any(x in dt for x in ["Zamalek", "الزمالك"]):
            dist = "الزمالك"
            gov = "القاهرة"
        elif any(x in dt for x in ["Dokki", "دقي"]):
            dist = "الدقي"
            gov = "الجيزة"
        elif any(x in dt for x in ["Mohandessin", "مهندسين"]):
            dist = "المهندسين"
            gov = "الجيزة"
        elif any(x in dt for x in ["October", "أكتوبر", "6th"]):
            dist = "مدينة 6 أكتوبر"
            gov = "الجيزة"
        elif any(x in dt for x in ["Zayed", "زايد"]):
            dist = "الشيخ زايد"
            gov = "الجيزة"
        elif any(x in dt for x in ["Tagamoa", "New Cairo", "التجمع", "القاهرة الجديدة"]):
            dist = "التجمع الخامس"
            gov = "القاهرة"

    return gov, dist

def get_features_and_services(main_cat, sub_cat):
    feats = ["يقبل الدفع بالبطاقات الائتمانية", "مناسب للمجموعات والعائلات"]
    services = ["دفع بالفيزا", "ساحة انتظار"]
    
    if main_cat == "أكل ومشروبات":
        feats.extend(["شبكة واي فاي مجانية", "خيارات نباتية متوفرة", "أماكن عائلية وكابلز"])
        services.append("توصيل طلبات")
    elif main_cat == "صحة":
        feats.extend(["مداخل سهلة للكراسي المتحركة", "مرافق مريحة للزوار", "مناسب لجميع الأعمار"])
        services.append("طوارئ 24 ساعة")
    elif main_cat == "تسوق":
        feats.extend(["مداخل سهلة للكراسي المتحركة", "مرافق مريحة للزوار", "مناسب لجميع الأعمار"])
        services.extend(["توصيل طلبات", "خدمة عملاء"])
    elif main_cat == "إقامة وسياحة":
        feats.extend(["شبكة واي فاي مجانية", "مرافق مريحة للزوار", "أماكن هادئة"])
        services.extend(["خدمة غرف", "استقبال 24 ساعة", "مطعم"])
    elif main_cat == "ترفيه":
        feats.extend(["مناسب للأطفال والعائلات", "مناسب لجميع الأعمار", "مرافق مريحة للزوار"])
        services.append("حجز تذاكر مسبق")
    elif main_cat == "رياضة":
        feats.extend(["مرافق مريحة للزوار", "مدربين معتمدين"])
        services.append("اشتراكات شهرية وسنوية")
    elif main_cat == "تعليم":
        feats.extend(["شبكة واي فاي مجانية", "مداخل سهلة للكراسي المتحركة", "مرافق مريحة للزوار"])
        services.append("شؤون طلاب وتدريب")
    else:
        feats.extend(["مرافق مريحة للزوار", "مناسب لجميع الأعمار"])

    return ", ".join(feats), ", ".join(services)

def get_working_hours(main_cat, sub_cat):
    if main_cat == "صحة" and sub_cat in ["مستشفيات", "صيدليات", "إسعاف"]:
        return {
            "sun": "24/7", "mon": "24/7", "tue": "24/7", "wed": "24/7",
            "thu": "24/7", "fri": "24/7", "sat": "24/7"
        }
    elif main_cat == "إقامة وسياحة" and sub_cat in ["فنادق"]:
        return {
            "sun": "24/7", "mon": "24/7", "tue": "24/7", "wed": "24/7",
            "thu": "24/7", "fri": "24/7", "sat": "24/7"
        }
    elif main_cat == "أكل ومشروبات":
        return {
            "sun": "09:00 ص - 12:00 ص", "mon": "09:00 ص - 12:00 ص", "tue": "09:00 ص - 12:00 ص",
            "wed": "09:00 ص - 12:00 ص", "thu": "09:00 ص - 01:00 ص", "fri": "01:00 م - 01:00 ص",
            "sat": "09:00 ص - 12:00 ص"
        }
    elif main_cat in ["تسوق", "ترفيه"]:
        return {
            "sun": "10:00 ص - 11:00 م", "mon": "10:00 ص - 11:00 م", "tue": "10:00 ص - 11:00 م",
            "wed": "10:00 ص - 11:00 م", "thu": "10:00 ص - 12:00 ص", "fri": "01:00 م - 12:00 ص",
            "sat": "10:00 ص - 11:00 م"
        }
    elif main_cat in ["خدمات حكومية", "خدمات مالية"]:
        return {
            "sun": "08:30 ص - 03:00 م", "mon": "08:30 ص - 03:00 م", "tue": "08:30 ص - 03:00 م",
            "wed": "08:30 ص - 03:00 م", "thu": "08:30 ص - 03:00 م", "fri": "إجازة",
            "sat": "إجازة"
        }
    elif main_cat == "تعليم":
        return {
            "sun": "08:00 ص - 04:00 م", "mon": "08:00 ص - 04:00 م", "tue": "08:00 ص - 04:00 م",
            "wed": "08:00 ص - 04:00 م", "thu": "08:00 ص - 04:00 م", "fri": "إجازة",
            "sat": "09:00 ص - 02:00 م"
        }
    else:
        return {
            "sun": "09:00 ص - 10:00 م", "mon": "09:00 ص - 10:00 م", "tue": "09:00 ص - 10:00 م",
            "wed": "09:00 ص - 10:00 م", "thu": "09:00 ص - 11:00 م", "fri": "02:00 م - 11:00 م",
            "sat": "09:00 ص - 10:00 م"
        }

def build_admin_import_excel():
    places_data = []
    seen = set()

    for item in raw_data:
        tags = item.get("tags", {})
        if not tags:
            continue
            
        name_ar, name_en = format_names(tags)
        if not name_ar or len(name_ar.strip()) < 2:
            continue

        lat, lon = determine_coords(item)
        if not lat or not lon:
            continue

        key = (name_ar.lower(), round(lat, 3), round(lon, 3))
        if key in seen:
            continue
        seen.add(key)

        main_cat, sub_cat, place_type, icon = map_to_admin_category(tags)
        gov, dist = determine_district_and_gov(lat, lon, tags)

        street = tags.get("addr:street", "").strip()
        housenumber = tags.get("addr:housenumber", "").strip()
        full_addr = tags.get("addr:full", "").strip()
        if not full_addr:
            if housenumber and street:
                full_addr = f"{housenumber} شارع {street}، {dist}، {gov}"
            elif street:
                full_addr = f"شارع {street}، {dist}، {gov}"
            else:
                full_addr = f"منطقة {dist}، {gov}"

        gmaps_url = f"https://maps.google.com/?q={lat},{lon}"
        phone = tags.get("phone") or tags.get("contact:phone") or tags.get("contact:mobile") or ""
        website = tags.get("website") or tags.get("contact:website") or ""

        wh = get_working_hours(main_cat, sub_cat)
        feats, services = get_features_and_services(main_cat, sub_cat)

        short_desc = f"{place_type} مميز في {dist}"
        detailed_desc = f"{name_ar} - يقع في {full_addr} ويقدم أفضل الخدمات للزوار والعملاء."

        row_dict = {
            "الاسم": name_ar,
            "الاسم (بالإنجليزية)": name_en,
            "القسم الرئيسي": main_cat,
            "الأقسام الفرعية": sub_cat,
            "المدينة / المنطقة": dist,
            "العنوان بالتفصيل": full_addr,
            "رابط جوجل ماب": gmaps_url,
            "المحافظة": gov,
            "الهواتف": phone,
            "مواعيد الأحد": wh["sun"],
            "مواعيد الإثنين": wh["mon"],
            "مواعيد الثلاثاء": wh["tue"],
            "مواعيد الأربعاء": wh["wed"],
            "مواعيد الخميس": wh["thu"],
            "مواعيد الجمعة": wh["fri"],
            "مواعيد السبت": wh["sat"],
            "مواعيد العمل": "",
            "معلومات مفيدة (المميزات)": feats,
            "خط العرض": lat,
            "خط الطول": lon,
            "وصف قصير": short_desc,
            "الوصف التفصيلي": detailed_desc,
            "رابط الصورة الرئيسية": "",
            "روابط المنيو": "",
            "موقع الويب": website,
            "الخدمات": services,
            "نوع المكان": place_type,
            "أيقونة النوع": icon
        }
        places_data.append(row_dict)

    print(f"Total structured places: {len(places_data)}")

    wb = openpyxl.Workbook()
    ws_places = wb.active
    ws_places.title = "الأماكن"
    ws_places.views.sheetView[0].rightToLeft = True

    headers = [
        "الاسم", "الاسم (بالإنجليزية)", "القسم الرئيسي", "الأقسام الفرعية", "المدينة / المنطقة", "العنوان بالتفصيل", "رابط جوجل ماب", "المحافظة", "الهواتف",
        "مواعيد الأحد", "مواعيد الإثنين", "مواعيد الثلاثاء", "مواعيد الأربعاء", "مواعيد الخميس", "مواعيد الجمعة", "مواعيد السبت",
        "مواعيد العمل",
        "معلومات مفيدة (المميزات)",
        "خط العرض", "خط الطول", "وصف قصير", "الوصف التفصيلي", "رابط الصورة الرئيسية", "روابط المنيو",
        "موقع الويب", "الخدمات", "نوع المكان", "أيقونة النوع"
    ]

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    regular_font = Font(name="Segoe UI", size=10, color="0F172A")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    ws_places.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, start=1):
        c = ws_places.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        col_letter = get_column_letter(col_idx)
        ws_places.column_dimensions[col_letter].width = max(len(h) * 2 + 4, 16)

    ws_places.column_dimensions["A"].width = 30 # الاسم العربي
    ws_places.column_dimensions["B"].width = 28 # الاسم الإنجليزي
    ws_places.column_dimensions["F"].width = 36 # العنوان
    ws_places.column_dimensions["G"].width = 36 # رابط جوجل ماب
    ws_places.column_dimensions["R"].width = 40 # المميزات
    ws_places.column_dimensions["V"].width = 45 # الوصف التفصيلي

    ws_places.freeze_panes = "A2"

    for row_idx, p in enumerate(places_data, start=2):
        ws_places.row_dimensions[row_idx].height = 20
        is_alt = (row_idx % 2 == 0)
        for col_idx, h in enumerate(headers, start=1):
            val = p.get(h, "")
            c = ws_places.cell(row=row_idx, column=col_idx, value=val)
            c.font = regular_font
            c.border = cell_border
            if is_alt:
                c.fill = alt_fill
            if h in ["خط العرض", "خط الطول", "المحافظة", "المدينة / المنطقة", "مواعيد الأحد", "مواعيد الإثنين", "مواعيد الثلاثاء", "مواعيد الأربعاء", "مواعيد الخميس", "مواعيد الجمعة", "مواعيد السبت"]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif h == "الاسم (بالإنجليزية)":
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")

    # Sheet 2: دليل التصنيفات الفرعية
    ws_guide = wb.create_sheet(title="دليل التصنيفات الفرعية")
    ws_guide.views.sheetView[0].rightToLeft = True
    guide_headers = ["القسم الرئيسي", "التصنيفات الفرعية المتاحة (يفصل بينها بفصلة)"]
    ws_guide.append(guide_headers)
    
    categories_guide = [
        ("أكل ومشروبات", "مطاعم ، فاست فود ، مخابز وحلويات ، كافيهات ، عصائر"),
        ("صحة", "مستشفيات ، عيادات ، صيدليات ، عيادات أسنان ، مراكز عيون ، معامل تحاليل ، مراكز أشعة ، إسعاف"),
        ("تسوق", "مولات ، سوبر ماركت ، محلات موبايلات ، محلات كمبيوتر ، ملابس ، أحذية ، مجوهرات ، مستحضرات تجميل ، أثاث ، مكتبات ، ألعاب ، مستلزمات حيوانات ، ورد وهدايا ، أدوات منزلية"),
        ("سيارات", "محطات بنزين ، مراكز صيانة ، معارض سيارات ، كاوتش ، مغاسل سيارات ، مواقف سيارات"),
        ("إقامة وسياحة", "فنادق ، شقق فندقية ، بيوت ضيافة ، مخيمات ، شركات سياحة"),
        ("ترفيه", "سينما ، ملاهي ، أكوا بارك ، بولينج ، حمامات سباحة ، مسارح ، متاحف ، معارض ، فعاليات"),
        ("رياضة", "جيم ، ملاعب ، ملاعب تنس ، أكاديميات رياضية ، تأجير دراجات"),
        ("خدمات حكومية", "مصالح حكومية ، قسم شرطة ، مطافي ، محاكم ، بريد ، شهر عقاري"),
        ("خدمات مالية", "بنوك ، ماكينات ATM ، صرافة"),
        ("أماكن دينية", "مساجد ، كنائس"),
        ("تعليم", "مدارس ، جامعات ، مراكز تعليم ، مراكز لغات ، حضانات"),
        ("أعمال", "شركات ، مكاتب ، مكاتب محاماة ، مكاتب محاسبة ، شركات عقارات"),
        ("خدمات", "مغاسل ، صالونات حلاقة ، بيوتي سنتر ، مفاتيح ، سباك ، كهربائي ، تكييف ، شركات شحن ، نقل أثاث"),
        ("أماكن عامة", "حدائق ، شواطئ ، محميات ، ميادين")
    ]
    for main_lbl, subs in categories_guide:
        ws_guide.append([main_lbl, subs])

    for col_idx in [1, 2]:
        c = ws_guide.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.column_dimensions["A"].width = 25
    ws_guide.column_dimensions["B"].width = 80

    # Sheet 3: دليل مواعيد العمل اليومية
    ws_wh = wb.create_sheet(title="دليل مواعيد العمل اليومية")
    ws_wh.views.sheetView[0].rightToLeft = True
    wh_headers = ["اسم عمود اليوم في Excel", "مثال التوقيت المدخل في الخلية", "شرح النتيجة"]
    ws_wh.append(wh_headers)
    wh_samples = [
        ["مواعيد الأحد (أو الأحد)", "09:00 ص - 11:00 م", "يفتح الساعة 9 صباحاً ويغلق 11 مساءً يوم الأحد"],
        ["مواعيد الخميس (أو الخميس)", "09:00 ص - 12:00 م", "يفتح الساعة 9 صباحاً ويغلق 12 منتصف الليل يوم الخميس"],
        ["مواعيد الجمعة (أو الجمعة)", "إجازة (أو مغلق)", "يوم الجمعة عطلة رسمية للمكان"],
        ["مواعيد الأحد إلى السبت", "24/7", "المكان يعمل 24 ساعة طوال اليوم"]
    ]
    for r in wh_samples:
        ws_wh.append(r)
    for col_idx in [1, 2, 3]:
        c = ws_wh.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_wh.column_dimensions["A"].width = 30
    ws_wh.column_dimensions["B"].width = 30
    ws_wh.column_dimensions["C"].width = 50

    # Sheet 4: دليل المميزات والمعلومات
    ws_feat = wb.create_sheet(title="دليل المميزات والمعلومات")
    ws_feat.views.sheetView[0].rightToLeft = True
    feat_headers = ["الأيقونة", "معلومات مفيدة / الميزة", "طريقة الكتابة في Excel (يمكن اختيار أكثر من ميزة بفصلة)"]
    ws_feat.append(feat_headers)
    feat_samples = [
        ["bx bx-wifi", "شبكة واي فاي مجانية", "شبكة واي فاي مجانية (أو free_wifi)"],
        ["bx bx-credit-card", "يقبل الدفع بالبطاقات الائتمانية", "يقبل الدفع بالبطاقات الائتمانية (أو accepts_credit_cards)"],
        ["bx bx-group", "مناسب للمجموعات والعائلات", "مناسب للمجموعات والعائلات (أو suitable_for_groups)"],
        ["bx bx-wheelchair", "مداخل سهلة للكراسي المتحركة", "مداخل سهلة للكراسي المتحركة (أو wheelchair_accessible)"],
        ["bx bx-leaf", "خيارات نباتية متوفرة", "خيارات نباتية متوفرة (أو vegetarian_options)"],
        ["bx bx-smile", "مناسب لجميع الأعمار", "مناسب لجميع الأعمار (أو suitable_for_all_ages)"],
        ["bx bx-heart", "أماكن عائلية وكابلز", "أماكن عائلية وكابلز (أو family_friendly)"],
        ["bx bx-moon", "أماكن هادئة", "أماكن هادئة (أو quiet_place)"],
        ["bx bx-check-circle", "مرافق مريحة للزوار", "مرافق مريحة للزوار (أو comfortable_facilities)"]
    ]
    for r in feat_samples:
        ws_feat.append(r)
    for col_idx in [1, 2, 3]:
        c = ws_feat.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_feat.column_dimensions["A"].width = 20
    ws_feat.column_dimensions["B"].width = 35
    ws_feat.column_dimensions["C"].width = 65

    # Top 100 version
    wb_100 = openpyxl.Workbook()
    ws_100 = wb_100.active
    ws_100.title = "الأماكن"
    ws_100.views.sheetView[0].rightToLeft = True
    ws_100.row_dimensions[1].height = 28
    for col_idx, h in enumerate(headers, start=1):
        c = ws_100.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        col_letter = get_column_letter(col_idx)
        ws_100.column_dimensions[col_letter].width = max(len(h) * 2 + 4, 16)
    ws_100.column_dimensions["A"].width = 30
    ws_100.column_dimensions["B"].width = 28
    ws_100.column_dimensions["F"].width = 36
    ws_100.column_dimensions["G"].width = 36
    ws_100.column_dimensions["R"].width = 40
    ws_100.column_dimensions["V"].width = 45
    ws_100.freeze_panes = "A2"

    for row_idx, p in enumerate(places_data[:100], start=2):
        ws_100.row_dimensions[row_idx].height = 20
        is_alt = (row_idx % 2 == 0)
        for col_idx, h in enumerate(headers, start=1):
            val = p.get(h, "")
            c = ws_100.cell(row=row_idx, column=col_idx, value=val)
            c.font = regular_font
            c.border = cell_border
            if is_alt:
                c.fill = alt_fill
            if h in ["خط العرض", "خط الطول", "المحافظة", "المدينة / المنطقة", "مواعيد الأحد", "مواعيد الإثنين", "مواعيد الثلاثاء", "مواعيد الأربعاء", "مواعيد الخميس", "مواعيد الجمعة", "مواعيد السبت"]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif h == "الاسم (بالإنجليزية)":
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")

    file_3500 = "places_template_cairo_giza_3500.xlsx"
    file_100 = "places_template_cairo_giza_100.xlsx"
    
    wb.save(file_3500)
    wb_100.save(file_100)

    desktop = os.path.expanduser("~/Desktop")
    if os.path.exists(desktop):
        for fname in [file_3500, file_100]:
            target_path = os.path.join(desktop, fname)
            try:
                shutil.copy2(fname, target_path)
                print(f"Copied {fname} to Desktop")
            except Exception as e:
                alt_target = os.path.join(desktop, f"new_{fname}")
                try:
                    shutil.copy2(fname, alt_target)
                    print(f"File was busy, copied as {alt_target}")
                except Exception as e2:
                    print(f"Could not copy to Desktop: {e2}")

    print("Generation complete!")

if __name__ == "__main__":
    build_admin_import_excel()
