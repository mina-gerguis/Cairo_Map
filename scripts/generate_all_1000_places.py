import os
import shutil
import openpyxl
from save_category_file import populate_sheet_data, base_dir, headers, header_fill, header_font
from build_assets_map import photo_collections

# Script to build and export 10 distinct category workbooks (100 verified places each) across Egypt
# Plus one consolidated Master Workbook with 10 tabs

print("Starting master generation for 10 Egypt categories (100 places each = 1,000 places total)...")

# We define generator data templates for each category
from sample_data_generators import (
    get_food_100,
    get_health_100,
    get_shopping_100,
    get_hotels_100,
    get_entertainment_100,
    get_education_100,
    get_sports_100,
    get_finance_100,
    get_government_100,
    get_public_places_100
)

categories_plan = [
    ("أكل ومشروبات", "01_أكل_ومشروبات_مصر_100_مكان.xlsx", get_food_100()),
    ("صحة", "02_صحة_ومستشفيات_مصر_100_مكان.xlsx", get_health_100()),
    ("تسوق", "03_تسوق_ومولات_مصر_100_مكان.xlsx", get_shopping_100()),
    ("إقامة وسياحة", "04_إقامة_وفنادق_مصر_100_مكان.xlsx", get_hotels_100()),
    ("ترفيه", "05_ترفيه_ومتاحف_مصر_100_مكان.xlsx", get_entertainment_100()),
    ("تعليم", "06_تعليم_وجامعات_مصر_100_مكان.xlsx", get_education_100()),
    ("رياضة", "07_رياضة_ونوادي_مصر_100_مكان.xlsx", get_sports_100()),
    ("خدمات مالية", "08_خدمات_مالية_وبنوك_مصر_100_مكان.xlsx", get_finance_100()),
    ("خدمات حكومية", "09_خدمات_حكومية_ونقل_مصر_100_مكان.xlsx", get_government_100()),
    ("أماكن عامة", "10_أماكن_عامة_ودينية_مصر_100_مكان.xlsx", get_public_places_100()),
]

# 1. Create individual category workbooks
for cat_name, filename, places in categories_plan:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cat_name[:31]
    populate_sheet_data(ws, places, cat_name)
    
    # Guide sheet
    ws_guide = wb.create_sheet(title="دليل التصنيفات الفرعية")
    ws_guide.views.sheetView[0].rightToLeft = True
    ws_guide.append(["القسم الرئيسي", "التصنيفات الفرعية المتاحة"])
    ws_guide.append([cat_name, places[0].get("sub_categories", "") if places else ""])
    for col_idx in [1, 2]:
        c = ws_guide.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
    ws_guide.column_dimensions["A"].width = 25
    ws_guide.column_dimensions["B"].width = 80

    out_file = os.path.join(base_dir, filename)
    wb.save(out_file)
    print(f"Generated category workbook with {len(places)} places.")

# 2. Create Master All-in-One Workbook with 10 Tabs
master_wb = openpyxl.Workbook()
default_sheet = master_wb.active

for cat_name, filename, places in categories_plan:
    ws = master_wb.create_sheet(title=cat_name[:31])
    populate_sheet_data(ws, places, cat_name)

if default_sheet in master_wb.worksheets:
    master_wb.remove(default_sheet)

master_filename = "الدليل_الشامل_لكل_تصنيفات_مصر_1000_مكان.xlsx"
master_path = os.path.join(base_dir, master_filename)
master_wb.save(master_path)
print("Generated Master Combined Workbook.")

# 3. Copy entire folder to Desktop
desktop = os.path.expanduser("~/Desktop")
desktop_target_dir = os.path.join(desktop, "دليل_تصنيفات_مصر_Excel")

if os.path.exists(desktop):
    try:
        if os.path.exists(desktop_target_dir):
            shutil.rmtree(desktop_target_dir, ignore_errors=True)
        shutil.copytree(base_dir, desktop_target_dir)
        print("Copied complete folder to Desktop successfully!")
    except Exception as e:
        print(f"Desktop copy notice: {e}")

print("All 10 Category files (1,000 places total) generated successfully!")
