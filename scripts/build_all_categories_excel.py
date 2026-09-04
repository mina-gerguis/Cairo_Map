import os
import shutil
import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create output directories
base_dir = "دليل_تصنيفات_مصر_Excel"
os.makedirs(base_dir, exist_ok=True)

headers = [
    "الاسم", "الاسم (بالإنجليزية)", "القسم الرئيسي", "الأقسام الفرعية", "المدينة / المنطقة", "العنوان بالتفصيل", "رابط جوجل ماب", "المحافظة", "الهواتف",
    "مواعيد الأحد", "مواعيد الإثنين", "مواعيد الثلاثاء", "مواعيد الأربعاء", "مواعيد الخميس", "مواعيد الجمعة", "مواعيد السبت",
    "مواعيد العمل",
    "معلومات مفيدة (المميزات)",
    "خط العرض", "خط الطول", "وصف قصير", "الوصف التفصيلي", "رابط الصورة الرئيسية", "روابط المنيو",
    "موقع الويب", "الخدمات", "نوع المكان", "أيقونة النوع"
]

header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
cell_border = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)
regular_font = Font(name="Segoe UI", size=10, color="0F172A")
link_font = Font(name="Segoe UI", size=10, color="2563EB", underline="single")
alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

def format_and_save_sheet(places_list, category_name, filename_ar, filename_en):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = category_name[:31]
    ws.views.sheetView[0].rightToLeft = True

    ws.row_dimensions[1].height = 30
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(h) * 2 + 4, 18)

    ws.column_dimensions["A"].width = 34 # الاسم
    ws.column_dimensions["B"].width = 32 # الاسم بالإنجليزي
    ws.column_dimensions["D"].width = 28 # الأقسام الفرعية
    ws.column_dimensions["E"].width = 22 # المدينة
    ws.column_dimensions["F"].width = 42 # العنوان
    ws.column_dimensions["G"].width = 38 # جوجل ماب
    ws.column_dimensions["R"].width = 45 # المميزات
    ws.column_dimensions["V"].width = 55 # الوصف التفصيلي
    ws.column_dimensions["W"].width = 48 # رابط الصورة الرئيسية
    ws.column_dimensions["X"].width = 48 # روابط المنيو
    ws.column_dimensions["Y"].width = 32 # موقع الويب

    ws.freeze_panes = "A2"

    for row_idx, p in enumerate(places_list, start=2):
        ws.row_dimensions[row_idx].height = 24
        is_alt = (row_idx % 2 == 0)
        
        row_data = [
            p.get("name", ""),
            p.get("name_en", ""),
            p.get("category", category_name),
            p.get("sub_categories", ""),
            p.get("city", ""),
            p.get("full_address", ""),
            p.get("google_maps_url", ""),
            p.get("governorate", "القاهرة"),
            p.get("phones", ""),
            p.get("wh_sun", "09:00 ص - 10:00 م"),
            p.get("wh_mon", "09:00 ص - 10:00 م"),
            p.get("wh_tue", "09:00 ص - 10:00 م"),
            p.get("wh_wed", "09:00 ص - 10:00 م"),
            p.get("wh_thu", "09:00 ص - 11:00 م"),
            p.get("wh_fri", "01:00 م - 11:00 م"),
            p.get("wh_sat", "09:00 ص - 10:00 م"),
            p.get("working_hours", ""),
            p.get("features", "شبكة واي فاي مجانية, يقبل الدفع بالبطاقات الائتمانية, مناسب للمجموعات والعائلات"),
            p.get("lat", 30.0444),
            p.get("lon", 31.2357),
            p.get("short_desc", ""),
            p.get("detailed_desc", ""),
            p.get("image_url", ""),
            p.get("menu_images", ""),
            p.get("website", ""),
            p.get("services", "دفع بالفيزا, ساحة انتظار"),
            p.get("place_type", ""),
            p.get("icon", "bx-store")
        ]
        
        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = regular_font
            c.border = cell_border
            if is_alt:
                c.fill = alt_fill
                
            h_name = headers[col_idx - 1]
            if h_name in ["خط العرض", "خط الطول", "المحافظة", "المدينة / المنطقة", "مواعيد الأحد", "مواعيد الإثنين", "مواعيد الثلاثاء", "مواعيد الأربعاء", "مواعيد الخميس", "مواعيد الجمعة", "مواعيد السبت", "الهواتف", "أيقونة النوع"]:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif h_name in ["الاسم (بالإنجليزية)", "رابط جوجل ماب", "رابط الصورة الرئيسية", "روابط المنيو", "موقع الويب"]:
                c.alignment = Alignment(horizontal="left", vertical="center")
                if str(val).startswith("http"):
                    c.font = link_font
            else:
                c.alignment = Alignment(horizontal="right", vertical="center")

    # Sheet 2: Guide
    ws_guide = wb.create_sheet(title="دليل التصنيفات الفرعية")
    ws_guide.views.sheetView[0].rightToLeft = True
    ws_guide.append(["القسم الرئيسي", "التصنيفات الفرعية المتاحة (يفصل بينها بفصلة)"])
    ws_guide.append([category_name, p.get("sub_categories", "")])
    for col_idx in [1, 2]:
        c = ws_guide.cell(row=1, column=col_idx)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws_guide.column_dimensions["A"].width = 25
    ws_guide.column_dimensions["B"].width = 80

    path_ar = os.path.join(base_dir, filename_ar)
    path_en = os.path.join(base_dir, filename_en)
    wb.save(path_ar)
    wb.save(path_en)

print("Setup helper initialized.")
